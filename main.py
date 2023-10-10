# 샘플 Python 스크립트입니다.

# ⌃R을(를) 눌러 실행하거나 내 코드로 바꿉니다.
# 클래스, 파일, 도구 창, 액션 및 설정을 어디서나 검색하려면 ⇧ 두 번을(를) 누릅니다.

import pymysql
from openpyxl import load_workbook
import pandas as pd
from tabulate import tabulate
from flask import Flask, render_template, request
import timer

sensor_db = pymysql.connect(
    user='root',
    passwd='12341234',
    host='127.0.0.1',
    db='D2_69',
    charset='utf8'
)

#cursor = sensor_db.cursor(pymysql.cursors.DictCursor)
cursor = sensor_db.cursor()

def season_result (season):

    load_wb = ""
    load_ws= ""
    table_name = ""

    if season == 2:
        load_wb = load_workbook("input/시즌2.xlsx", read_only=True, data_only=True)
        print(load_wb.sheetnames)
        load_ws = load_wb['결과']
        table_name = "D2_69.table_s2_result"

    if season == 3:
        load_wb = load_workbook("input/시즌3.xlsx", read_only=True, data_only=True)
        print(load_wb.sheetnames)
        load_ws = load_wb['결과']
        table_name = "D2_69.table_s3_result"

    index = []
    get_cells = load_ws['A2': 'U2']
    for row in get_cells:
        for cell in row:
            index.append(cell.value)
    print(index)

    rank = []
    get_cells = load_ws['A3': 'A22']
    for row in get_cells:
        for cell in row:
            rank.append(cell.value)
    print(rank)

    tier = []
    get_cells = load_ws['B3': 'B22']
    for row in get_cells:
        for cell in row:
            tier.append(cell.value)
    print(tier)

    id = []
    get_cells = load_ws['C3': 'C22']
    for row in get_cells:
        for cell in row:
            id.append(cell.value)
    print(id)

    tier_score = []
    get_cells = load_ws['D3': 'D22']
    for row in get_cells:
        for cell in row:
            tier_score.append(round(cell.value, 2))
    print(tier_score)

    mmr_rank = []
    get_cells = load_ws['E3': 'E22']
    for row in get_cells:
        for cell in row:
            mmr_rank.append(cell.value)
    print(mmr_rank)

    total_game = []
    get_cells = load_ws['F3': 'F22']
    for row in get_cells:
        for cell in row:
            total_game.append(cell.value)
    print(total_game)

    total_win = []
    get_cells = load_ws['G3': 'G22']
    for row in get_cells:
        for cell in row:
            total_win.append(cell.value)
    print(total_win)

    total_lose = []
    get_cells = load_ws['H3': 'H22']
    for row in get_cells:
        for cell in row:
            total_lose.append(cell.value)
    print(total_lose)

    total_rate = []
    get_cells = load_ws['I3': 'I22']
    for row in get_cells:
        for cell in row:
            total_rate.append(round(cell.value*100, 1))
    print(total_rate)

    druid_win = []
    get_cells = load_ws['J3': 'J22']
    for row in get_cells:
        for cell in row:
            druid_win.append(cell.value)
    print(druid_win)

    druid_lose = []
    get_cells = load_ws['K3': 'K22']
    for row in get_cells:
        for cell in row:
            druid_lose.append(cell.value)
    print(druid_lose)

    druid_rate = []
    get_cells = load_ws['L3': 'L22']
    for row in get_cells:
        for cell in row:
            druid_rate.append(round(cell.value*100, 1))
    print(druid_rate)

    ass_win = []
    get_cells = load_ws['M3': 'M22']
    for row in get_cells:
        for cell in row:
            ass_win.append(cell.value)
    print(ass_win)

    ass_lose = []
    get_cells = load_ws['N3': 'N22']
    for row in get_cells:
        for cell in row:
            ass_lose.append(cell.value)
    print(ass_lose)

    ass_rate = []
    get_cells = load_ws['O3': 'O22']
    for row in get_cells:
        for cell in row:
            ass_rate.append(round(cell.value*100, 1))
    print(ass_rate)

    nec_win = []
    get_cells = load_ws['P3': 'P22']
    for row in get_cells:
        for cell in row:
            nec_win.append(cell.value)
    print(nec_win)

    nec_lose = []
    get_cells = load_ws['Q3': 'Q22']
    for row in get_cells:
        for cell in row:
            nec_lose.append(cell.value)
    print(nec_lose)

    nec_rate = []
    get_cells = load_ws['R3': 'R22']
    for row in get_cells:
        for cell in row:
            nec_rate.append(round(cell.value*100, 1))
    print(nec_rate)

    pala_win = []
    get_cells = load_ws['S3': 'S22']
    for row in get_cells:
        for cell in row:
            pala_win.append(cell.value)
    print(pala_win)

    pala_lose = []
    get_cells = load_ws['T3': 'T22']
    for row in get_cells:
        for cell in row:
            pala_lose.append(cell.value)
    print(pala_lose)

    pala_rate = []
    get_cells = load_ws['U3': 'U22']
    for row in get_cells:
        for cell in row:
            pala_rate.append(round(cell.value*100, 1))
    print(pala_rate)

    i = 0
    index = 1
    for i in range(len(rank)):
        sql = ("INSERT INTO " + table_name + " VALUES(" + '"' + str(index) + '","' +
                  rank[i] + '","' +
                  tier[i] + '","' +
                  id[i] + '","' +
                  str(tier_score[i]) + '","' +
                  str(mmr_rank[i]) + '","' +
                  str(total_game[i]) + '","' +
                  str(total_win[i]) + '","' +
                  str(total_lose[i]) + '","' +
                  str(total_rate[i]) + "%" + '","' +
                  str(druid_win[i]) + '","' +
                  str(druid_lose[i]) + '","' +
                  str(druid_rate[i]) + "%" + '","' +
                  str(ass_win[i]) + '","' +
                  str(ass_lose[i]) + '","' +
                  str(ass_rate[i]) + "%" + '","' +
                  str(nec_win[i]) + '","' +
                  str(nec_lose[i]) + '","' +
                  str(nec_rate[i]) + "%" + '","' +
                  str(pala_win[i]) + '","' +
                  str(pala_lose[i]) + '","' +
                  str(pala_rate[i]) + "%" + '")')
        cursor.execute(sql)
        index = index + 1

    load_wb.close()

def season_char (season):

    load_wb = ""
    load_ws = ""
    table_name = ""

    if season == 2:
        load_wb = load_workbook("input/시즌2.xlsx", read_only=True, data_only=True)
        print(load_wb.sheetnames)
        load_ws = load_wb['주캐릭']
        table_name = "D2_69.table_s2_char"

    if season == 3:
        load_wb = load_workbook("input/시즌3.xlsx", read_only=True, data_only=True)
        print(load_wb.sheetnames)
        load_ws = load_wb['주캐릭']
        table_name = "D2_69.table_s3_char"

    id = []
    get_cells = load_ws['A3': 'A39']
    for row in get_cells:
        for cell in row:
            id.append(cell.value)
    print(id)

    dru = []
    get_cells = load_ws['B3': 'B39']
    for row in get_cells:
        for cell in row:
            dru.append(cell.value)
    print(dru)

    ass = []
    get_cells = load_ws['C3': 'C39']
    for row in get_cells:
        for cell in row:
            ass.append(cell.value)
    print(ass)

    nec = []
    get_cells = load_ws['D3': 'D39']
    for row in get_cells:
        for cell in row:
            nec.append(cell.value)
    print(nec)

    pala = []
    get_cells = load_ws['E3': 'E39']
    for row in get_cells:
        for cell in row:
            pala.append(cell.value)
    print(pala)

    for i in range(len(id)):
        sql = "INSERT INTO " + table_name + " VALUES(" + '"' + id[i] + '","' + dru[i] + '","' + ass[i] + '","' + nec[i] + '","' + pala[i] + '")'
        cursor.execute(sql)


def season_player (season):

    load_wb = ""
    load_ws = ""
    table_name_duo_win_rate = ""
    table_name_relative_record = ""
    table_name_win_lose_straight = ""
    season_player = []
    player_name = ""

    if season == 2:
        load_wb = load_workbook("input/시즌2.xlsx", read_only=True, data_only=True)
        print(load_wb.sheetnames)
        load_ws = load_wb['린스']
        table_name_duo_win_rate = "d2_69.table_s2_duo_win_rate_"
        table_name_relative_record = "d2_69.table_s2_relative_record_"
        table_name_win_lose_straight = "d2_69.table_s2_win_lose_straight_"

    if season == 3:
        load_wb = load_workbook("input/시즌3.xlsx", read_only=True, data_only=True)
        print(load_wb.sheetnames)
        load_ws = load_wb['주캐릭']
        table_name_duo_win_rate = "d2_69.table_s3_duo_win_rate_"
        table_name_relative_record = "d2_69.table_s3_relative_record_"
        table_name_win_lose_straight = "d2_69.table_s3_win_lose_straight_"

    season_player = load_wb.sheetnames
    del season_player[0]
    del season_player[0]
    for i in season_player:
        if i == "린스":
            player_name = "lince"
        elif i == "태식":
            player_name = "taesik"
        elif i == "반짝":
            player_name = "banzzak"
        elif i == "민형":
            player_name = "mh"
        elif i == "하리보":
            player_name = "haribo"
        elif i == "야로":
            player_name = "yaro"
        elif i == "훈":
            player_name = "hoon"
        elif i == "넘버":
            player_name = "number"
        elif i == "용이":
            player_name = "yongi"
        elif i == "용갈":
            player_name = "yongal"
        elif i == "울프":
            player_name = "wolf"
        elif i == "조네":
            player_name = "jone"
        elif i == "사카시":
            player_name = "sakasy"
        elif i == "블핑":
            player_name = "blackpink"
        elif i == "도건":
            player_name = "dogeon"
        elif i == "살수":
            player_name = "salsu"
        elif i == "참치":
            player_name = "chamchi"
        elif i == "르기":
            player_name = "rgi"
        elif i == "플토":
            player_name = "pluto"
        elif i == "필찌":
            player_name = "pilzzi"
        elif i == "솔":
            player_name = "sol"
        elif i == "라위":
            player_name = "lawe"
        else:
            player_name = ""

        load_ws = load_wb[i]

        duo_player1 = []
        get_cells = load_ws['A3': 'A22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    duo_player1.append(cell.value)
        print(duo_player1)

        duo_player2 = []
        get_cells = load_ws['B3': 'B22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    duo_player2.append(cell.value)
        print(duo_player2)

        duo_win = []
        get_cells = load_ws['C3': 'C22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    duo_win.append(cell.value)
        print(duo_win)

        duo_lose = []
        get_cells = load_ws['D3': 'D22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    duo_lose.append(cell.value)
        print(duo_lose)

        duo_rate = []
        get_cells = load_ws['E3': 'E22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    duo_rate.append(round(cell.value * 100, 2))
        print(duo_rate)

        relative_player1 = []
        get_cells = load_ws['G3': 'G22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    relative_player1.append(cell.value)
        print(relative_player1)

        relative_vs1 = []
        get_cells = load_ws['H3': 'H22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    relative_vs1.append(cell.value)
        print(relative_vs1)

        relative_vs2 = []
        get_cells = load_ws['I3': 'I22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    relative_vs2.append(cell.value)
        print(relative_vs2)

        relative_player2 = []
        get_cells = load_ws['J3': 'J22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    relative_player2.append(cell.value)
        print(relative_player2)

        win_straight = []
        get_cells = load_ws['M3': 'M22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    win_straight.append(cell.value)
        print(win_straight)

        lose_straight = []
        get_cells = load_ws['N3': 'N22']
        for row in get_cells:
            for cell in row:
                if (cell.value is not None):
                    lose_straight.append(cell.value)
        print(lose_straight)

        sql = "DELETE FROM " + table_name_duo_win_rate + player_name
        cursor.execute(sql)

        sql = "DELETE FROM " + table_name_relative_record + player_name
        cursor.execute(sql)

        sql = "DELETE FROM " + table_name_win_lose_straight + player_name
        cursor.execute(sql)

        for i in range(len(duo_player1)):
            if (i is not None):
                sql = "INSERT INTO " + table_name_duo_win_rate + player_name + " VALUES(" + '"' + duo_player1[i] + '","' + duo_player2[i] + '","' + str(duo_win[i]) + '","' + str(duo_lose[i]) + '","' + str(duo_rate[i]) + "%" + '")'
                cursor.execute(sql)

        for i in range(len(relative_player1)):
            if (i is not None):
                sql = "INSERT INTO " + table_name_relative_record + player_name + " VALUES(" + '"' + relative_player1[i] + '","' + str(relative_vs1[i]) + '","' + str(relative_vs2[i]) + '","' + relative_player2[i] + '")'
                cursor.execute(sql)

        for i in range(len(win_straight)):
            if (i is not None):
                sql = "INSERT INTO " + table_name_win_lose_straight + player_name + " VALUES(" + '"' + str(win_straight[i]) + '","' + str(lose_straight[i]) + '")'
                cursor.execute(sql)


# 스크립트를 실행하려면 여백의 녹색 버튼을 누릅니다.
if __name__ == '__main__':

    load_wb = load_workbook("input/69내전기록표_0920_v27.xlsm", read_only=True, data_only=True)
    print(load_wb.sheetnames)

    load_ws = load_wb['현재시즌']

    #####################################################

    index = []
    get_cells = load_ws['B83': 'G83']
    for row in get_cells:
        for cell in row:
            index.append(cell.value)
    print(index)

    rank = []
    get_cells = load_ws['B84': 'B103']
    for row in get_cells:
        for cell in row:
            rank.append(cell.value)
    print(rank)

    tier = []
    get_cells = load_ws['C84': 'C103']
    for row in get_cells:
        for cell in row:
            tier.append(cell.value)
    print(tier)

    id = []
    get_cells = load_ws['D84': 'D103']
    for row in get_cells:
        for cell in row:
            id.append(cell.value)
    print(id)

    tier_score = []
    get_cells = load_ws['E84': 'E103']
    for row in get_cells:
        for cell in row:
            tier_score.append(round(cell.value, 2))
    print(tier_score)

    mmr_rank = []
    get_cells = load_ws['F84': 'F103']
    for row in get_cells:
        for cell in row:
            mmr_rank.append(cell.value)
    print(mmr_rank)

    total_game = []
    get_cells = load_ws['G84': 'G103']
    for row in get_cells:
        for cell in row:
            total_game.append(cell.value)
    print(total_game)

    #result = pd.DataFrame({'전체 순위':rank, '티어':tier, '아이디':id, '티어점수':tier_score, 'MMR순위':mmr_rank, '총 게임수':total_game})
    #print(tabulate(result, headers='keys', tablefmt='grid', showindex=True, stralign='center'))

    i = 0
    index = 1

    sql = "DELETE FROM D2_69.table_cur_s_result"
    cursor.execute(sql)

    for i in range(len(rank)):
        sql = ("INSERT INTO D2_69.table_cur_s_result VALUES(" + '"' + str(index) + '","' +
               rank[i] + '","' +
               tier[i] + '","' +
               id[i] + '","' +
               str(tier_score[i]) + '","' +
               str(mmr_rank[i]) + '","' +
               str(total_game[i]) + '")')

        cursor.execute(sql)
        index = index + 1

    sql = "DELETE FROM D2_69.table_s2_char"
    cursor.execute(sql)
    sql = "DELETE FROM D2_69.table_s3_char"
    cursor.execute(sql)

    sql = "DELETE FROM D2_69.table_s2_result"
    cursor.execute(sql)

    sql = "DELETE FROM D2_69.table_s3_result"
    cursor.execute(sql)

    sql = "DELETE FROM d2_69.table_s2_duo_win_rate_lince"
    cursor.execute(sql)

    season_result(2)
    season_result(3)

    season_char(2)
    season_char(3)

    season_player(2)
    season_player(3)

    cursor.close()
    sensor_db.commit()